import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
warnings.filterwarnings('ignore')

# ── Load raw data ──────────────────────────────────────────────────────────────
xl = pd.read_excel('/mnt/user-data/uploads/fazaaaexcel.xlsx', sheet_name=None)
orders   = xl['Orders'].copy()
products = xl['Products'].copy()
customers= xl['Customers'].copy()

# ── Clean & enrich orders ──────────────────────────────────────────────────────
orders['Revenue'] = orders['Quantity'] * orders['Products.Price (INR)']
orders['Order_Date_Real'] = pd.to_datetime(orders['Order_Date'])
orders['Month_Num'] = orders['Order_Date_Real'].dt.month

MONTH_ORDER = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']

# ── KPI calculations ───────────────────────────────────────────────────────────
total_revenue     = orders['Revenue'].sum()
total_orders      = len(orders)
avg_order_value   = total_revenue / total_orders
avg_delivery_days = orders['diff_order_delivery'].mean()
total_customers   = customers['Customer_ID'].nunique()
total_products    = products['Product_ID'].nunique()

# Revenue by occasion
rev_by_occasion = orders.groupby('Occasion')['Revenue'].sum().sort_values(ascending=False)

# Revenue by month
rev_by_month = orders.groupby('Month Name')['Revenue'].sum()
rev_by_month = rev_by_month.reindex([m for m in MONTH_ORDER if m in rev_by_month.index])

# Orders by day
orders_by_day = orders.groupby('Day Name(Order Date)')['Order_ID'].count()
day_order = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
orders_by_day = orders_by_day.reindex([d for d in day_order if d in orders_by_day.index])

# Top cities
top_cities = orders.groupby('Location').agg(
    Orders=('Order_ID','count'),
    Revenue=('Revenue','sum')
).sort_values('Revenue', ascending=False).head(10)

# Category performance
cat_perf = orders.merge(products[['Product_ID','Category']], on='Product_ID', how='left')
cat_rev  = cat_perf.groupby('Category')['Revenue'].sum().sort_values(ascending=False)

# Peak order hours
hour_orders = orders.groupby('Hour(Order Time)')['Order_ID'].count().sort_index()

# Gender split (customers)
gender_split = customers['Gender'].value_counts()

# Product performance
prod_perf = orders.merge(products[['Product_ID','Product_Name','Category']], on='Product_ID', how='left')
top_products = prod_perf.groupby('Product_Name').agg(
    Orders=('Order_ID','count'),
    Revenue=('Revenue','sum'),
    Qty=('Quantity','sum')
).sort_values('Revenue', ascending=False).head(10)

# ── Colour palette ─────────────────────────────────────────────────────────────
DARK_BG    = "1A1A2E"
MID_BG     = "16213E"
ACCENT1    = "0F3460"
ACCENT2    = "E94560"
ACCENT3    = "533483"
GOLD       = "F5A623"
GREEN      = "27AE60"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
HEADER_TXT = "FFFFFF"
SUBHDR     = "B0C4DE"
CARD_BG    = "0D2137"

def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mk_font(bold=False, size=11, color=WHITE, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def mk_border(color="D0D0D0", thin=True):
    s = "thin" if thin else "medium"
    side = Side(style=s, color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def mk_align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Create workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 – EXECUTIVE DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("📊 Dashboard")
ws.sheet_properties.tabColor = "E94560"
ws.freeze_panes = "A5"

# Set column widths
col_widths = [2,14,14,14,14,14,14,14,2]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Row heights
for r in range(1, 80):
    ws.row_dimensions[r].height = 18
ws.row_dimensions[1].height = 8
ws.row_dimensions[2].height = 42
ws.row_dimensions[3].height = 8
ws.row_dimensions[4].height = 24
ws.row_dimensions[5].height = 8
for r in [6,7,8,9,10,11]:
    ws.row_dimensions[r].height = 22

# Background entire sheet
for row in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=9):
    for cell in row:
        cell.fill = mk_fill(DARK_BG)

# ── Title bar ──────────────────────────────────────────────────────────────────
ws.merge_cells("B2:H2")
title_cell = ws["B2"]
title_cell.value = "🎁  GiftMart India — Executive Sales Dashboard"
title_cell.font  = Font(name="Calibri", bold=True, size=22, color=WHITE)
title_cell.alignment = mk_align("left")
title_cell.fill = mk_fill(ACCENT1)

# Subtitle
ws.merge_cells("B3:H3")
sub = ws["B3"]
sub.value = "FY Analysis  |  1,000 Orders  |  100 Customers  |  70 Products  |  12 Cities"
sub.font  = mk_font(size=10, color=SUBHDR, italic=True)
sub.alignment = mk_align("left")
sub.fill = mk_fill(ACCENT1)

# ── Section header helper ──────────────────────────────────────────────────────
def section_header(ws, row, col, text, colspan=1, end_col=None):
    if end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font  = mk_font(bold=True, size=11, color=WHITE)
    c.fill  = mk_fill(ACCENT2)
    c.alignment = mk_align("left")

# ── KPI Cards (row 5) ─────────────────────────────────────────────────────────
section_header(ws, 4, 2, "  KEY PERFORMANCE INDICATORS", end_col=8)

kpis = [
    ("💰 Total Revenue", f"₹{total_revenue/1e6:.2f}M", GREEN),
    ("📦 Total Orders",  f"{total_orders:,}",           GOLD),
    ("🛒 Avg Order Value",f"₹{avg_order_value:,.0f}",   ACCENT3),
    ("🚚 Avg Delivery",  f"{avg_delivery_days:.1f} days", ACCENT2),
    ("👥 Customers",     f"{total_customers}",           ACCENT1),
    ("🏷️ Products",      f"{total_products}",            "2C7BB6"),
    ("🏙️ Cities",        f"{orders['Location'].nunique()}", "8E44AD"),
]

kpi_cols = [2,3,4,5,6,7,8]
for (label, value, color), col in zip(kpis, kpi_cols):
    # Card rows 6–10
    for r in range(6, 11):
        c = ws.cell(row=r, column=col)
        c.fill = mk_fill(CARD_BG)
        if r == 7:
            c.value = label
            c.font  = mk_font(size=9, color=SUBHDR)
            c.alignment = mk_align()
        elif r == 8:
            c.value = value
            c.font  = Font(name="Calibri", bold=True, size=16, color=color)
            c.alignment = mk_align()
        elif r == 6 or r == 10:
            c.fill = mk_fill(color)
    # top accent line
    ws.row_dimensions[6].height = 5
    ws.row_dimensions[10].height = 5

# ── Analytics tables (start row 12) ───────────────────────────────────────────
# ── Revenue by Occasion ────────────────────────────────────────────────────────
section_header(ws, 12, 2, "  Revenue by Occasion", end_col=4)
section_header(ws, 12, 5, "  Revenue by Category", end_col=8)

ws.row_dimensions[12].height = 22

# Occasion table
hdrs = ["Occasion", "Revenue (₹)", "Orders", "Share %"]
occ_data = orders.groupby('Occasion').agg(
    Revenue=('Revenue','sum'),
    Orders=('Order_ID','count')
).sort_values('Revenue', ascending=False).reset_index()
occ_data['Share'] = occ_data['Revenue'] / occ_data['Revenue'].sum() * 100

for ci, h in enumerate(hdrs, 2):
    c = ws.cell(row=13, column=ci)
    c.value = h
    c.font  = mk_font(bold=True, size=9, color=WHITE)
    c.fill  = mk_fill(ACCENT1)
    c.alignment = mk_align()
    c.border = mk_border(ACCENT1)

for ri, row_d in occ_data.iterrows():
    r = 14 + ri
    ws.row_dimensions[r].height = 20
    fill_c = mk_fill("0A1628") if ri % 2 == 0 else mk_fill(CARD_BG)
    vals = [row_d['Occasion'], f"₹{row_d['Revenue']:,.0f}",
            row_d['Orders'], f"{row_d['Share']:.1f}%"]
    for ci, v in enumerate(vals, 2):
        c = ws.cell(row=r, column=ci)
        c.value = v
        c.font  = mk_font(size=9)
        c.fill  = fill_c
        c.alignment = mk_align()
        c.border = mk_border("1A2A3A")

# Category table
cat_data = cat_perf.groupby('Category').agg(
    Revenue=('Revenue','sum'),
    Orders=('Order_ID','count')
).sort_values('Revenue', ascending=False).reset_index()
cat_data['Avg'] = cat_data['Revenue'] / cat_data['Orders']

hdrs2 = ["Category", "Revenue (₹)", "Orders", "Avg (₹)"]
for ci, h in enumerate(hdrs2, 5):
    c = ws.cell(row=13, column=ci)
    c.value = h
    c.font  = mk_font(bold=True, size=9, color=WHITE)
    c.fill  = mk_fill(ACCENT1)
    c.alignment = mk_align()

for ri, row_d in cat_data.iterrows():
    r = 14 + ri
    ws.row_dimensions[r].height = 20
    fill_c = mk_fill("0A1628") if ri % 2 == 0 else mk_fill(CARD_BG)
    vals = [row_d['Category'], f"₹{row_d['Revenue']:,.0f}",
            row_d['Orders'], f"₹{row_d['Avg']:,.0f}"]
    for ci, v in enumerate(vals, 5):
        c = ws.cell(row=r, column=ci)
        c.value = v
        c.font  = mk_font(size=9)
        c.fill  = fill_c
        c.alignment = mk_align()
        c.border = mk_border("1A2A3A")

# ── Top 10 Cities ─────────────────────────────────────────────────────────────
base_r = 14 + len(occ_data) + 2
section_header(ws, base_r, 2, "  Top 10 Cities by Revenue", end_col=4)
section_header(ws, base_r, 5, "  Day-wise Order Pattern", end_col=8)
ws.row_dimensions[base_r].height = 22

city_hdrs = ["City", "Revenue (₹)", "Orders"]
for ci, h in enumerate(city_hdrs, 2):
    c = ws.cell(row=base_r+1, column=ci)
    c.value = h; c.font = mk_font(bold=True, size=9)
    c.fill = mk_fill(ACCENT1); c.alignment = mk_align()

for ri, (city, row_d) in enumerate(top_cities.iterrows()):
    r = base_r + 2 + ri
    ws.row_dimensions[r].height = 19
    fill_c = mk_fill("0A1628") if ri % 2 == 0 else mk_fill(CARD_BG)
    for ci, v in enumerate([city, f"₹{row_d['Revenue']:,.0f}", row_d['Orders']], 2):
        c = ws.cell(row=r, column=ci)
        c.value = v; c.font = mk_font(size=9)
        c.fill = fill_c; c.alignment = mk_align()

# Day-wise table
day_hdrs = ["Day", "Orders", "% of Total"]
for ci, h in enumerate(day_hdrs, 5):
    c = ws.cell(row=base_r+1, column=ci)
    c.value = h; c.font = mk_font(bold=True, size=9)
    c.fill = mk_fill(ACCENT1); c.alignment = mk_align()

for ri, (day, cnt) in enumerate(orders_by_day.items()):
    r = base_r + 2 + ri
    ws.row_dimensions[r].height = 19
    fill_c = mk_fill("0A1628") if ri % 2 == 0 else mk_fill(CARD_BG)
    pct = cnt / total_orders * 100
    for ci, v in enumerate([day, cnt, f"{pct:.1f}%"], 5):
        c = ws.cell(row=r, column=ci)
        c.value = v; c.font = mk_font(size=9)
        c.fill = fill_c; c.alignment = mk_align()


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 – CLEANED DATA (Orders)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📋 Orders Data")
ws2.sheet_properties.tabColor = "0F3460"
ws2.freeze_panes = "A2"

orders_clean = orders[['Order_ID','Customer_ID','Product_ID','Quantity',
                        'Location','Occasion','Month Name','Day Name(Order Date)',
                        'Hour(Order Time)','diff_order_delivery',
                        'Products.Price (INR)','Revenue']].copy()
orders_clean.columns = ['Order ID','Customer ID','Product ID','Quantity',
                        'City','Occasion','Month','Day','Order Hour',
                        'Delivery Days','Unit Price (₹)','Revenue (₹)']

col_ws = [8,12,10,9,18,18,12,12,11,14,14,14]
headers = list(orders_clean.columns)

for i, (h, w) in enumerate(zip(headers, col_ws), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    c = ws2.cell(row=1, column=i)
    c.value = h
    c.font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill  = mk_fill(ACCENT1)
    c.alignment = mk_align()
    c.border = mk_border(DARK_BG)

for ri, row_vals in enumerate(orders_clean.values, 2):
    fill_c = mk_fill(LIGHT_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for ci, v in enumerate(row_vals, 1):
        c = ws2.cell(row=ri, column=ci)
        c.value = float(v) if isinstance(v, (np.integer, np.floating)) else v
        c.font  = Font(name="Calibri", size=9)
        c.fill  = fill_c
        c.alignment = mk_align("center")
        c.border = mk_border("D0D0D0")

# Conditional format Revenue column (col 12)
ws2.conditional_formatting.add(
    f"L2:L{len(orders_clean)+1}",
    DataBarRule(start_type='min', end_type='max',
                color=ACCENT2)
)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 – PRODUCTS & CUSTOMERS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🛍️ Products")
ws3.sheet_properties.tabColor = "533483"
ws3.freeze_panes = "A2"

prod_display = products[['Product_ID','Product_Name','Category','Price (INR)','Occasion']].copy()
prod_hdrs = ['Product ID','Product Name','Category','Price (₹)','Occasion']
prod_widths = [10,22,16,12,18]

for i, (h, w) in enumerate(zip(prod_hdrs, prod_widths), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
    c = ws3.cell(row=1, column=i)
    c.value = h; c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = mk_fill(ACCENT3); c.alignment = mk_align(); c.border = mk_border(DARK_BG)

for ri, row_vals in enumerate(prod_display.values, 2):
    fill_c = mk_fill(LIGHT_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for ci, v in enumerate(row_vals, 1):
        c = ws3.cell(row=ri, column=ci)
        c.value = float(v) if isinstance(v, (np.integer, np.floating)) else v
        c.font = Font(name="Calibri", size=9)
        c.fill = fill_c; c.alignment = mk_align("center"); c.border = mk_border("D0D0D0")

# Price data bar
ws3.conditional_formatting.add(
    f"D2:D{len(prod_display)+1}",
    DataBarRule(start_type='min', end_type='max', color="27AE60")
)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 4 – MONTHLY TREND ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📈 Monthly Analysis")
ws4.sheet_properties.tabColor = GOLD

month_agg = orders.groupby('Month Name').agg(
    Revenue=('Revenue','sum'),
    Orders=('Order_ID','count'),
    Avg_Value=('Revenue','mean'),
    Avg_Delivery=('diff_order_delivery','mean'),
    Qty=('Quantity','sum')
).reindex([m for m in MONTH_ORDER if m in orders['Month Name'].unique()]).reset_index()

month_hdrs = ['Month','Revenue (₹)','Orders','Avg Order Value (₹)','Avg Delivery Days','Total Qty Sold']
month_widths = [14,16,10,20,18,16]

for i, (h, w) in enumerate(zip(month_hdrs, month_widths), 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
    c = ws4.cell(row=1, column=i)
    c.value = h; c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = mk_fill(ACCENT1); c.alignment = mk_align(); c.border = mk_border(DARK_BG)

for ri, row_vals in enumerate(month_agg.values, 2):
    fill_c = mk_fill(LIGHT_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for ci, v in enumerate(row_vals, 1):
        c = ws4.cell(row=ri, column=ci)
        c.value = round(float(v), 2) if isinstance(v, (np.integer, np.floating)) else v
        c.font = Font(name="Calibri", size=9)
        c.fill = fill_c; c.alignment = mk_align("center"); c.border = mk_border("D0D0D0")

# Revenue color scale
ws4.conditional_formatting.add(
    f"B2:B{len(month_agg)+1}",
    ColorScaleRule(start_type='min', start_color='FFF7DC',
                   end_type='max', end_color=ACCENT2)
)

# Bar chart – Monthly Revenue
bar = BarChart()
bar.type = "col"; bar.grouping = "clustered"
bar.title = "Monthly Revenue (₹)"; bar.style = 10
bar.y_axis.title = "Revenue (₹)"; bar.x_axis.title = "Month"
bar.width = 22; bar.height = 12
bar.shape = 4

data_ref = Reference(ws4, min_col=2, min_row=1, max_row=len(month_agg)+1)
cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=len(month_agg)+1)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.series[0].graphicalProperties.solidFill = ACCENT2
ws4.add_chart(bar, "H2")

# Line chart – Orders trend
line = LineChart()
line.title = "Monthly Order Volume"; line.style = 10
line.y_axis.title = "Orders"; line.x_axis.title = "Month"
line.width = 22; line.height = 12
ord_ref = Reference(ws4, min_col=3, min_row=1, max_row=len(month_agg)+1)
line.add_data(ord_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.series[0].graphicalProperties.line.solidFill = GOLD
line.series[0].graphicalProperties.line.width = 28000
ws4.add_chart(line, "H22")

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 5 – OCCASION & PRODUCT PERFORMANCE
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🎉 Occasion Analysis")
ws5.sheet_properties.tabColor = ACCENT2

occ_full = orders.groupby('Occasion').agg(
    Revenue=('Revenue','sum'),
    Orders=('Order_ID','count'),
    Avg_Value=('Revenue','mean'),
    Avg_Delivery=('diff_order_delivery','mean'),
    Total_Qty=('Quantity','sum')
).sort_values('Revenue', ascending=False).reset_index()
occ_full['Revenue Share %'] = (occ_full['Revenue'] / occ_full['Revenue'].sum() * 100).round(1)

occ_hdrs = ['Occasion','Revenue (₹)','Orders','Avg Order Value (₹)','Avg Delivery Days','Total Qty','Revenue Share %']
occ_widths = [18,16,10,20,18,12,16]

for i, (h, w) in enumerate(zip(occ_hdrs, occ_widths), 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
    c = ws5.cell(row=1, column=i)
    c.value = h; c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = mk_fill(ACCENT2); c.alignment = mk_align(); c.border = mk_border(DARK_BG)

for ri, row_vals in enumerate(occ_full.values, 2):
    fill_c = mk_fill(LIGHT_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for ci, v in enumerate(row_vals, 1):
        c = ws5.cell(row=ri, column=ci)
        c.value = round(float(v), 2) if isinstance(v, (np.integer, np.floating)) else v
        c.font = Font(name="Calibri", size=9)
        c.fill = fill_c; c.alignment = mk_align("center"); c.border = mk_border("D0D0D0")

# Pie chart – Revenue share by occasion
pie = PieChart()
pie.title = "Revenue Share by Occasion"; pie.style = 10
pie.width = 18; pie.height = 14
pie_data = Reference(ws5, min_col=2, min_row=1, max_row=len(occ_full)+1)
pie_labels = Reference(ws5, min_col=1, min_row=2, max_row=len(occ_full)+1)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_labels)
pie.dataLabels = DataLabelList(showPercent=True, showSerName=False, showCatName=True)
ws5.add_chart(pie, "I2")

# Top products table
top_p = top_products.reset_index()
tp_start = len(occ_full) + 4
ws5.cell(row=tp_start, column=1).value = "TOP 10 PRODUCTS BY REVENUE"
ws5.cell(row=tp_start, column=1).font = Font(name="Calibri", bold=True, size=11, color=ACCENT1)

tp_hdrs = ['Product Name','Orders','Revenue (₹)','Qty Sold']
tp_widths = [24,10,16,12]
for i, (h, w) in enumerate(zip(tp_hdrs, tp_widths), 1):
    ws5.column_dimensions[get_column_letter(i)].width = max(ws5.column_dimensions[get_column_letter(i)].width, w)
    c = ws5.cell(row=tp_start+1, column=i)
    c.value = h; c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    c.fill = mk_fill(ACCENT1); c.alignment = mk_align(); c.border = mk_border(DARK_BG)

for ri, row_vals in enumerate(top_p.values, tp_start+2):
    fill_c = mk_fill(LIGHT_GRAY) if (ri - tp_start) % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for ci, v in enumerate(row_vals, 1):
        c = ws5.cell(row=ri, column=ci)
        c.value = round(float(v), 2) if isinstance(v, (np.integer, np.floating)) else v
        c.font = Font(name="Calibri", size=9)
        c.fill = fill_c; c.alignment = mk_align("center"); c.border = mk_border("D0D0D0")

# Revenue data bar for top products
ws5.conditional_formatting.add(
    f"C{tp_start+2}:C{tp_start+1+len(top_p)}",
    DataBarRule(start_type='min', end_type='max', color=GREEN)
)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 6 – CUSTOMER ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("👥 Customers")
ws6.sheet_properties.tabColor = GREEN

cust_clean = customers[['Customer_ID','Name','City','Gender']].copy()
cust_orders = orders.groupby('Customer_ID').agg(
    Total_Orders=('Order_ID','count'),
    Total_Revenue=('Revenue','sum'),
    Avg_Order=('Revenue','mean')
).reset_index()
cust_merged = cust_clean.merge(cust_orders, on='Customer_ID', how='left').fillna(0)
cust_merged = cust_merged.sort_values('Total_Revenue', ascending=False)

cust_hdrs = ['Customer ID','Name','City','Gender','Total Orders','Total Revenue (₹)','Avg Order (₹)']
cust_widths = [12,20,18,10,14,18,14]

for i, (h, w) in enumerate(zip(cust_hdrs, cust_widths), 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
    c = ws6.cell(row=1, column=i)
    c.value = h; c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = mk_fill(GREEN.replace("27AE60","1A8449")); c.alignment = mk_align(); c.border = mk_border(DARK_BG)

for ri, row_vals in enumerate(cust_merged.values, 2):
    fill_c = mk_fill(LIGHT_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for ci, v in enumerate(row_vals, 1):
        c = ws6.cell(row=ri, column=ci)
        c.value = round(float(v), 2) if isinstance(v, (np.integer, np.floating)) else v
        c.font = Font(name="Calibri", size=9)
        c.fill = fill_c; c.alignment = mk_align("center"); c.border = mk_border("D0D0D0")

ws6.conditional_formatting.add(
    f"F2:F{len(cust_merged)+1}",
    DataBarRule(start_type='min', end_type='max', color=ACCENT2)
)

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 7 – SUMMARY METRICS (for README reference)
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("📌 Summary Metrics")
ws7.sheet_properties.tabColor = ACCENT3

summary_data = [
    ("Metric", "Value", "Notes"),
    ("Total Orders", total_orders, "Across all occasions"),
    ("Total Revenue", f"₹{total_revenue:,.0f}", "Sum of Qty × Price"),
    ("Average Order Value", f"₹{avg_order_value:,.0f}", "Revenue / Orders"),
    ("Average Delivery Days", f"{avg_delivery_days:.1f}", "Order to delivery"),
    ("Top Occasion", rev_by_occasion.idxmax(), f"₹{rev_by_occasion.max():,.0f} revenue"),
    ("Top Month", rev_by_month.idxmax(), f"{rev_by_month.max():,.0f} revenue"),
    ("Top City", top_cities.index[0], f"₹{top_cities['Revenue'].iloc[0]:,.0f}"),
    ("Top Category", cat_rev.idxmax(), f"₹{cat_rev.max():,.0f}"),
    ("Peak Order Hour", f"{hour_orders.idxmax()}:00", f"{hour_orders.max()} orders"),
    ("Total Customers", total_customers, "Unique customers"),
    ("Total Products", total_products, "Unique SKUs"),
    ("Female Customers", f"{gender_split.get('Female',0)}", ""),
    ("Male Customers", f"{gender_split.get('Male',0)}", ""),
]

sm_widths = [28, 20, 30]
for i, w in enumerate(sm_widths, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

for ri, row_v in enumerate(summary_data, 1):
    for ci, v in enumerate(row_v, 1):
        c = ws7.cell(row=ri, column=ci)
        c.value = v
        if ri == 1:
            c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
            c.fill = mk_fill(ACCENT3)
        else:
            c.font = Font(name="Calibri", size=10)
            c.fill = mk_fill(LIGHT_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        c.alignment = mk_align("left")
        c.border = mk_border()

# ── Save ───────────────────────────────────────────────────────────────────────
out = "/home/claude/GiftMart_India_Sales_Dashboard.xlsx"
wb.save(out)
print("Saved:", out)
print(f"Sheets: {[s.title for s in wb.worksheets]}")
